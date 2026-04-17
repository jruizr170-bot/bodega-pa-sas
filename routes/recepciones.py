import os
import uuid
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func

import models, schemas
from database import get_db

router = APIRouter(prefix="/recepciones", tags=["recepciones"])

CLOUDINARY_URL = os.environ.get("CLOUDINARY_URL")
UPLOADS_BASE   = Path(__file__).parent.parent / "uploads"


# ── helpers ───────────────────────────────────────────────────────────────────

def _save_photo(file: UploadFile) -> str:
    """Sube foto a Cloudinary (si está configurado) o disco local. Devuelve URL."""
    raw = file.file.read()

    # Comprimir con Pillow si >1 MB
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(raw))
        img = img.convert("RGB")
        if max(img.size) > 2000:
            img.thumbnail((2000, 2000), Image.LANCZOS)
        buf = io.BytesIO()
        quality = 70 if len(raw) > 1_000_000 else 85
        img.save(buf, format="JPEG", quality=quality, optimize=True)
        raw = buf.getvalue()
    except Exception:
        pass

    if CLOUDINARY_URL:
        import cloudinary
        import cloudinary.uploader
        cloudinary.config(cloudinary_url=CLOUDINARY_URL)
        result = cloudinary.uploader.upload(
            raw,
            folder="bodega-pa-sas",
            resource_type="image",
        )
        return result["secure_url"]

    # Fallback local
    today = date.today()
    d = UPLOADS_BASE / str(today.year) / f"{today.month:02d}" / f"{today.day:02d}"
    d.mkdir(parents=True, exist_ok=True)
    nombre = f"{uuid.uuid4().hex}.jpg"
    (d / nombre).write_bytes(raw)
    return str((d / nombre).relative_to(UPLOADS_BASE.parent))


def _load_recepcion(recepcion_id: int, db: Session):
    rec = db.query(models.Recepcion).options(
        joinedload(models.Recepcion.bodega),
        joinedload(models.Recepcion.usuario),
        joinedload(models.Recepcion.proveedor),
        joinedload(models.Recepcion.items),
        joinedload(models.Recepcion.fotos),
    ).filter(models.Recepcion.id == recepcion_id).first()
    if not rec:
        raise HTTPException(status_code=404, detail="Recepción no encontrada")
    return rec


# ── CRUD ──────────────────────────────────────────────────────────────────────

@router.get("/", response_model=list[schemas.RecepcionOut])
def listar(
    skip: int = 0,
    limit: int = 50,
    bodega_id: Optional[int] = None,
    fecha: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(models.Recepcion).options(
        joinedload(models.Recepcion.bodega),
        joinedload(models.Recepcion.usuario),
        joinedload(models.Recepcion.proveedor),
        joinedload(models.Recepcion.items),
        joinedload(models.Recepcion.fotos),
    )
    if bodega_id:
        q = q.filter(models.Recepcion.bodega_id == bodega_id)
    if fecha:
        q = q.filter(func.date(models.Recepcion.fecha_registro) == fecha)
    q = q.order_by(models.Recepcion.fecha_registro.desc())
    return q.offset(skip).limit(limit).all()


@router.get("/{recepcion_id}", response_model=schemas.RecepcionOut)
def detalle(recepcion_id: int, db: Session = Depends(get_db)):
    return _load_recepcion(recepcion_id, db)


@router.post("/", response_model=schemas.RecepcionOut, status_code=201)
def crear(payload: schemas.RecepcionIn, db: Session = Depends(get_db)):
    # Resolver usuario
    usuario_id = payload.usuario_id
    if not usuario_id and payload.usuario_nombre:
        nombre = payload.usuario_nombre.strip()
        usr = db.query(models.Usuario).filter(
            models.Usuario.nombre.ilike(nombre)
        ).first()
        if not usr:
            usr = models.Usuario(nombre=nombre)
            db.add(usr)
            db.flush()
        usuario_id = usr.id

    # Resolver proveedor — si viene NIT y nombre y no hay proveedor_id, crear
    proveedor_id = payload.proveedor_id
    if not proveedor_id and payload.proveedor_nit and payload.proveedor_nombre:
        nit = payload.proveedor_nit.strip()
        prov = db.query(models.Proveedor).filter(
            models.Proveedor.nit == nit
        ).first()
        if not prov:
            prov = models.Proveedor(nit=nit, nombre=payload.proveedor_nombre.strip())
            db.add(prov)
            db.flush()
        proveedor_id = prov.id

    rec = models.Recepcion(
        fecha_factura=payload.fecha_factura,
        numero_factura=payload.numero_factura,
        proveedor_id=proveedor_id,
        proveedor_nombre=payload.proveedor_nombre if not proveedor_id else None,
        bodega_id=payload.bodega_id,
        usuario_id=usuario_id,
        observaciones=payload.observaciones,
        total_factura=payload.total_factura,
    )
    for it in payload.items:
        rec.items.append(models.RecepcionItem(**it.model_dump()))
    db.add(rec)
    db.commit()
    db.refresh(rec)
    return _load_recepcion(rec.id, db)


@router.post("/{recepcion_id}/foto", response_model=schemas.RecepcionOut)
def subir_foto(
    recepcion_id: int,
    foto: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    rec = db.query(models.Recepcion).filter(
        models.Recepcion.id == recepcion_id
    ).first()
    if not rec:
        raise HTTPException(status_code=404, detail="Recepción no encontrada")
    url = _save_photo(foto)
    db.add(models.RecepcionFoto(recepcion_id=recepcion_id, url=url))
    db.commit()
    return _load_recepcion(recepcion_id, db)


@router.delete("/{recepcion_id}", status_code=204)
def eliminar(recepcion_id: int, db: Session = Depends(get_db)):
    rec = db.query(models.Recepcion).filter(
        models.Recepcion.id == recepcion_id
    ).first()
    if not rec:
        raise HTTPException(status_code=404, detail="Recepción no encontrada")
    db.delete(rec)
    db.commit()


# ── Dashboard ─────────────────────────────────────────────────────────────────

@router.get("/dashboard/stats", response_model=schemas.DashboardStats)
def dashboard(db: Session = Depends(get_db)):
    hoy = datetime.utcnow().date()
    semana_pasada = hoy - timedelta(days=7)

    total      = db.query(func.count(models.Recepcion.id)).scalar() or 0
    hoy_count  = db.query(func.count(models.Recepcion.id)).filter(
        func.date(models.Recepcion.fecha_registro) == hoy).scalar() or 0
    sem_count  = db.query(func.count(models.Recepcion.id)).filter(
        func.date(models.Recepcion.fecha_registro) >= semana_pasada).scalar() or 0
    total_fact = db.query(func.sum(models.Recepcion.total_factura)).scalar() or 0.0

    por_bodega_raw = (
        db.query(models.Bodega.nombre, func.count(models.Recepcion.id).label("cantidad"))
        .join(models.Recepcion, models.Recepcion.bodega_id == models.Bodega.id, isouter=True)
        .group_by(models.Bodega.id).all()
    )

    ultimas = db.query(models.Recepcion).options(
        joinedload(models.Recepcion.bodega),
        joinedload(models.Recepcion.usuario),
        joinedload(models.Recepcion.proveedor),
        joinedload(models.Recepcion.items),
        joinedload(models.Recepcion.fotos),
    ).order_by(models.Recepcion.fecha_registro.desc()).limit(10).all()

    return schemas.DashboardStats(
        total_recepciones=total,
        recepciones_hoy=hoy_count,
        recepciones_semana=sem_count,
        total_facturado=total_fact,
        recepciones_por_bodega=[{"bodega": r[0], "cantidad": r[1]} for r in por_bodega_raw],
        ultimas_recepciones=ultimas,
    )


# ── Exportar Excel ────────────────────────────────────────────────────────────

@router.get("/exportar/excel")
def exportar_excel(bodega_id: Optional[int] = None, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    import io

    recs = listar(skip=0, limit=10000, bodega_id=bodega_id, db=db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Recepciones"

    h_fill = PatternFill("solid", fgColor="1E3A5F")
    h_font = Font(bold=True, color="FFFFFF")
    thin   = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))

    headers = ["ID","Fecha Registro","Fecha Factura","N° Factura",
               "Proveedor","NIT","Bodega","Usuario","Total Factura","Observaciones"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = h_fill; cell.font = h_font
        cell.alignment = Alignment(horizontal="center"); cell.border = thin

    alt = PatternFill("solid", fgColor="EBF0F8")
    for i, r in enumerate(recs):
        row = [
            r.id,
            r.fecha_registro.strftime("%Y-%m-%d %H:%M") if r.fecha_registro else "",
            r.fecha_factura or "",
            r.numero_factura or "",
            (r.proveedor.nombre if r.proveedor else r.proveedor_nombre) or "",
            (r.proveedor.nit if r.proveedor else "") or "",
            r.bodega.nombre if r.bodega else "",
            r.usuario.nombre if r.usuario else "",
            r.total_factura,
            r.observaciones or "",
        ]
        ws.append(row)
        fill = alt if i % 2 == 1 else None
        for cell in ws[i + 2]:
            if fill: cell.fill = fill
            cell.border = thin

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 40)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    nombre = f"recepciones_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre}"})
